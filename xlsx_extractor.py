import sys
from pathlib import Path
from typing import Any

VENDOR_DIR = Path(__file__).parent / ".vendor"
if VENDOR_DIR.exists():
    sys.path.insert(0, str(VENDOR_DIR))

import requests
from openpyxl import load_workbook


def extract_target_rows(source: dict[str, Any], targets: dict[str, Any], output_dir: Path) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_url = source.get("xlsx_url") or source["url"]
    xlsx_path = output_dir / f"{source['id']}.xlsx"

    response = requests.get(xlsx_url, timeout=60)
    response.raise_for_status()
    xlsx_path.write_bytes(response.content)

    target_ncms = {normalize_ncm(item["ncm"]) for item in targets["ncms"]}
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)

    rows: list[dict[str, Any]] = []
    texts: list[str] = []
    for sheet in workbook.worksheets:
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [clean_cell(value) for value in row]
            normalized = [normalize_ncm(value) for value in values]
            if not target_ncms.intersection(normalized):
                continue
            non_empty = [value for value in values if value]
            row_text = " | ".join(non_empty)
            rows.append({"sheet": sheet.title, "row": row_number, "values": non_empty})
            texts.append(f"TIPI XLSX {sheet.title} linha {row_number}: {row_text}")

    workbook.close()
    return {
        "url": source["url"],
        "final_url": xlsx_url,
        "status_code": response.status_code,
        "content_type": response.headers.get("content-type"),
        "downloaded_file": str(xlsx_path),
        "title": source["titulo"],
        "texts": texts,
        "text_blocks": [{"tag": "xlsx-row", "text": text} for text in texts],
        "rows": rows,
        "stats": {"total_texts": len(texts), "total_rows": len(rows)},
        "blocked": False,
        "blocked_reason": None,
        "success": True,
    }


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_ncm(value: str) -> str:
    return "".join(char for char in value if char.isdigit())
